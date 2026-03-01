from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, ConfigDict
from typing import Optional, List
import bcrypt
import secrets
from openpyxl import load_workbook
import io

# ---------------- ENV ----------------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.getenv("MONGO_URL")
DB_NAME = os.getenv("DB_NAME", "exam_db")

if not MONGO_URL:
    raise Exception("MONGO_URL not found in .env")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# ---------------- APP ----------------
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ⚠️ For production use Redis
sessions = {}

# ---------------- MODELS ----------------
class LoginRequest(BaseModel):
    enrollment_number: str
    password: str


class LoginResponse(BaseModel):
    success: bool
    message: str
    user: Optional[dict] = None
    session_token: Optional[str] = None


class StudentInfo(BaseModel):
    model_config = ConfigDict(extra="ignore")
    enrollment_number: str
    name: str
    branch: str
    role: str = "student"


class ExamDetails(BaseModel):
    model_config = ConfigDict(extra="ignore")
    enrollment_number: str
    exam_name: str
    room_number: str
    bench_number: str
    block: str
    exam_date: str


class DashboardResponse(BaseModel):
    student: StudentInfo
    exams: List[ExamDetails]


class AdminUploadResponse(BaseModel):
    success: bool
    message: str
    students_added: int
    exams_added: int


# ---------------- HELPERS ----------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())


def generate_session_token():
    return secrets.token_urlsafe(32)


async def get_current_user(request: Request):
    token = request.cookies.get("session_token")

    if not token or token not in sessions:
        raise HTTPException(status_code=401, detail="Not authenticated")

    return sessions[token]


# ---------------- AUTH ----------------
@api_router.post("/auth/login", response_model=LoginResponse)
async def login(data: LoginRequest, response: Response):

    user = await db.students.find_one(
        {"enrollment_number": data.enrollment_number},
        {"_id": 0}
    )

    if not user or not verify_password(data.password, user["password_hash"]):
        return LoginResponse(success=False, message="Invalid credentials")

    token = generate_session_token()

    sessions[token] = {
        "enrollment_number": user["enrollment_number"],
        "name": user["name"],
        "branch": user["branch"],
        "role": user["role"],
    }

    response.set_cookie(
        key="session_token",
        value=token,
        httponly=True,
        max_age=86400,
        samesite="lax",
    )

    return LoginResponse(
        success=True,
        message="Login successful",
        user=sessions[token],
        session_token=token,
    )


@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    token = request.cookies.get("session_token")

    if token:
        sessions.pop(token, None)

    response.delete_cookie("session_token")

    return {"success": True}


# ---------------- STUDENT ----------------
@api_router.get("/student/dashboard", response_model=DashboardResponse)
async def dashboard(request: Request):

    user = await get_current_user(request)

    student = await db.students.find_one(
        {"enrollment_number": user["enrollment_number"]},
        {"_id": 0, "password_hash": 0},
    )

    if not student:
        raise HTTPException(404, "Student not found")

    exams = await db.exams.find(
        {"enrollment_number": user["enrollment_number"]},
        {"_id": 0},
    ).to_list(100)

    return DashboardResponse(
        student=StudentInfo(**student),
        exams=[ExamDetails(**e) for e in exams],
    )


# ---------------- ADMIN UPLOAD ----------------
@api_router.post("/admin/upload-excel", response_model=AdminUploadResponse)
async def upload_excel(request: Request, file: UploadFile = File(...)):

    user = await get_current_user(request)

    if user["role"] != "admin":
        raise HTTPException(403, "Admin access required")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Upload Excel file")

    contents = await file.read()

    workbook = load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    students_added = 0
    exams_added = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row or not row[0]:
            continue

        (
            enrollment_number,
            name,
            branch,
            password,
            exam_name,
            room_number,
            bench_number,
            block,
            exam_date,
        ) = map(str, row[:9])

        exists = await db.students.find_one(
            {"enrollment_number": enrollment_number}
        )

        if not exists:
            await db.students.insert_one({
                "enrollment_number": enrollment_number,
                "name": name,
                "branch": branch,
                "password_hash": hash_password(password),
                "role": "student",
            })
            students_added += 1

        await db.exams.insert_one({
            "enrollment_number": enrollment_number,
            "exam_name": exam_name,
            "room_number": room_number,
            "bench_number": bench_number,
            "block": block,
            "exam_date": exam_date,
        })

        exams_added += 1

    return AdminUploadResponse(
        success=True,
        message="Upload successful",
        students_added=students_added,
        exams_added=exams_added,
    )


# ---------------- ROUTER ----------------
app.include_router(api_router)

# ---------------- CORS ----------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------- LOGGING ----------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ---------------- EVENTS ----------------
@app.on_event("shutdown")
async def shutdown():
    client.close()
